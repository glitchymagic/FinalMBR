import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, Series

# Create a new Excel workbook
def create_excel_report(filename):
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    return wb

# Function to style the header row
def style_header(ws, row_num=1):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

# Function to add borders to a range of cells
def add_borders(ws, cell_range):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

# Function to create the executive summary sheet
def create_executive_summary(wb, data):
    ws = wb.create_sheet("Executive Summary")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "East Region - June 2025 Executive Summary"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    ws['A3'] = "Total Incidents:"
    ws['B3'] = data['total_incidents']
    ws['A4'] = "Average MTTR (hours):"
    ws['B4'] = data['avg_mttr']
    ws['A5'] = "SLA Compliance:"
    ws['B5'] = f"{data['sla_compliance']}%"
    ws['A6'] = "First Call Resolution Rate:"
    ws['B6'] = f"{data['fcr_rate']}%"
    
    # Format numbers
    ws['B3'].number_format = '0'
    ws['B4'].number_format = '0.0'
    ws['B5'].number_format = '0.0%'
    ws['B6'].number_format = '0.0%'
    
    # Location breakdown
    ws['A8'] = "Location Breakdown"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Location"
    ws['B9'] = "Incidents"
    ws['C9'] = "MTTR (hours)"
    ws['D9'] = "FCR Rate"
    
    style_header(ws, 9)
    
    row = 10
    for loc in data['locations']:
        ws[f'A{row}'] = loc['name']
        ws[f'B{row}'] = loc['incidents']
        ws[f'C{row}'] = loc['mttr']
        ws[f'D{row}'] = loc['fcr_rate'] / 100
        
        ws[f'B{row}'].number_format = '0'
        ws[f'C{row}'].number_format = '0.0'
        ws[f'D{row}'].number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A9:D{row-1}')
    
    # Top technicians
    ws['A12'] = "Top Technicians"
    ws['A12'].font = Font(bold=True)
    
    ws['A13'] = "Technician"
    ws['B13'] = "Location"
    ws['C13'] = "Incidents"
    
    style_header(ws, 13)
    
    row = 14
    for tech in data['top_technicians']:
        ws[f'A{row}'] = tech['name']
        ws[f'B{row}'] = tech['location']
        ws[f'C{row}'] = tech['incidents']
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A13:C{row-1}')
    
    # Top KB articles
    ws['A17'] = "Top KB Articles"
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = "KB Article"
    ws['B18'] = "Description"
    ws['C18'] = "Usage Count"
    
    style_header(ws, 18)
    
    row = 19
    for kb in data['top_kb_articles']:
        ws[f'A{row}'] = kb['kb_id']
        ws[f'B{row}'] = kb['description']
        ws[f'C{row}'] = kb['count']
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A18:C{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    return ws

# Function to create a location-specific sheet
def create_location_sheet(wb, location_name, location_data):
    ws = wb.create_sheet(location_name)
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{location_name} - June 2025 Performance"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    ws['A3'] = "Total Incidents:"
    ws['B3'] = location_data['incidents']
    ws['A4'] = "Average MTTR (hours):"
    ws['B4'] = location_data['mttr']
    ws['A5'] = "SLA Compliance:"
    ws['B5'] = location_data['sla_compliance'] / 100
    ws['A6'] = "First Call Resolution Rate:"
    ws['B6'] = location_data['fcr_rate'] / 100
    
    # Format numbers
    ws['B3'].number_format = '0'
    ws['B4'].number_format = '0.0'
    ws['B5'].number_format = '0.0%'
    ws['B6'].number_format = '0.0%'
    
    # Technicians
    ws['A8'] = "Technicians"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Technician"
    ws['B9'] = "Incidents"
    ws['C9'] = "% of Location Total"
    
    style_header(ws, 9)
    
    row = 10
    if 'technicians' in location_data and location_data['technicians']:
        for tech in location_data['technicians']:
            ws[f'A{row}'] = tech['name']
            ws[f'B{row}'] = tech['incidents']
            ws[f'C{row}'] = tech['percentage'] / 100
            
            ws[f'B{row}'].number_format = '0'
            ws[f'C{row}'].number_format = '0.0%'
            
            row += 1
    else:
        ws[f'A{row}'] = "No technicians found for this location"
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
    
    # Add borders
    add_borders(ws, f'A9:C{row-1}')
    
    # KB Articles
    ws['A12'] = "KB Articles"
    ws['A12'].font = Font(bold=True)
    
    ws['A13'] = "KB Article"
    ws['B13'] = "Description"
    ws['C13'] = "Usage Count"
    
    style_header(ws, 13)
    
    row = 14
    if 'kb_articles' in location_data and location_data['kb_articles']:
        for kb in location_data['kb_articles']:
            ws[f'A{row}'] = kb['kb_id']
            ws[f'B{row}'] = kb['description']
            ws[f'C{row}'] = kb['count']
            
            row += 1
    else:
        ws[f'A{row}'] = "No KB articles found for this location"
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
    
    # Add borders
    add_borders(ws, f'A13:C{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    
    return ws

# Function to create the raw data sheet
def create_raw_data_sheet(wb, data):
    ws = wb.create_sheet("Raw Data")
    
    # Headers
    headers = ["Location", "Total Incidents", "MTTR (hours)", "SLA Compliance", "FCR Rate"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    style_header(ws)
    
    # Data
    row = 2
    for loc in data['locations']:
        ws.cell(row=row, column=1).value = loc['name']
        ws.cell(row=row, column=2).value = loc['incidents']
        ws.cell(row=row, column=3).value = loc['mttr']
        ws.cell(row=row, column=4).value = loc['sla_compliance'] / 100
        ws.cell(row=row, column=5).value = loc['fcr_rate'] / 100
        
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=5).number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A1:E{row-1}')
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return ws

# Function to create the incident details sheet
def create_incident_details_sheet(wb, data):
    ws = wb.create_sheet("Incident Details")
    
    # Headers
    headers = ["Location", "Technician", "Incidents", "KB Article", "Description", "Usage Count"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    style_header(ws)
    
    # Data
    row = 2
    for loc in data['locations']:
        if 'technicians' in loc and loc['technicians']:
            for tech in loc['technicians']:
                ws.cell(row=row, column=1).value = loc['name']
                ws.cell(row=row, column=2).value = tech['name']
                ws.cell(row=row, column=3).value = tech['incidents']
                row += 1
        else:
            ws.cell(row=row, column=1).value = loc['name']
            ws.cell(row=row, column=2).value = "No technicians found"
            row += 1
    
    # Add borders
    add_borders(ws, f'A1:F{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    
    return ws

# Main function to generate the report
def generate_east_report():
    # East Region data
    east_data = {
        'total_incidents': 136,  # 75 (Hoboken) + 61 (Herndon) + 0 (Charlotte)
        'avg_mttr': 2.0,  # Weighted average of Hoboken (3.7) and Herndon (0.1)
        'sla_compliance': 100.0,
        'fcr_rate': 99.3,  # Weighted average
        'locations': [
            {
                'name': 'Hoboken',
                'incidents': 75,
                'mttr': 3.7,
                'sla_compliance': 100.0,
                'fcr_rate': 98.7,
                'technicians': [
                    {'name': 'Harrison Lidoshore', 'incidents': 34, 'percentage': 45.3},
                    {'name': 'Angelique Smith', 'incidents': 22, 'percentage': 29.3},
                    {'name': 'Michael Jones', 'incidents': 19, 'percentage': 25.3}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 10},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 9},
                    {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 8},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 7},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 4}
                ]
            },
            {
                'name': 'Herndon',
                'incidents': 61,
                'mttr': 0.1,
                'sla_compliance': 100.0,
                'fcr_rate': 100.0,
                'technicians': [
                    {'name': 'Ali Nasrati', 'incidents': 32, 'percentage': 52.5},
                    {'name': 'Aadil Shafiq', 'incidents': 29, 'percentage': 47.5}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 9},
                    {'kb_id': 'KB1148665 v1.0', 'description': 'Hardware Diagnostics & Repair', 'count': 6},
                    {'kb_id': 'KB1148523 v4.0', 'description': 'Technical Support Issue', 'count': 3},
                    {'kb_id': 'KB0925519 v16.0', 'description': 'Zoom - Chat Issues | Zoom Chat turns off randomly in MyTech', 'count': 3},
                    {'kb_id': 'KB1148222 v3.0', 'description': 'User Account & Authentication Issues', 'count': 3}
                ]
            },
            {
                'name': 'Charlotte',
                'incidents': 0,
                'mttr': 0.0,
                'sla_compliance': 0.0,
                'fcr_rate': 0.0,
                'technicians': [],
                'kb_articles': []
            }
        ],
        'top_technicians': [
            {'name': 'Harrison Lidoshore', 'location': 'Hoboken', 'incidents': 34},
            {'name': 'Ali Nasrati', 'location': 'Herndon', 'incidents': 32},
            {'name': 'Aadil Shafiq', 'location': 'Herndon', 'incidents': 29},
            {'name': 'Angelique Smith', 'location': 'Hoboken', 'incidents': 22},
            {'name': 'Michael Jones', 'location': 'Hoboken', 'incidents': 19}
        ],
        'top_kb_articles': [
            {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 12},
            {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 9},
            {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 9},
            {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 8},
            {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 7}
        ]
    }
    
    # Create the workbook
    wb = create_excel_report("East_Region_June_2025_Report.xlsx")
    
    # Create sheets
    create_executive_summary(wb, east_data)
    
    for location in east_data['locations']:
        create_location_sheet(wb, location['name'], location)
    
    create_raw_data_sheet(wb, east_data)
    create_incident_details_sheet(wb, east_data)
    
    # Save the workbook
    wb.save("East_Region_June_2025_Report.xlsx")
    print("East Region report generated successfully!")

if __name__ == "__main__":
    generate_east_report() 