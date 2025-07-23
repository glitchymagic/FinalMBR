import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, Series

# Create a new Excel workbook
def create_excel_report(filename):
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    return wb

# Function to style the header row
def style_header(ws, row_num=1):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

# Function to add borders to a range of cells
def add_borders(ws, cell_range):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

# Function to create the executive summary sheet
def create_executive_summary(wb, data):
    ws = wb.create_sheet("Executive Summary")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "Central Region - June 2025 Executive Summary"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    ws['A3'] = "Total Incidents:"
    ws['B3'] = data['total_incidents']
    ws['A4'] = "Average MTTR (hours):"
    ws['B4'] = data['avg_mttr']
    ws['A5'] = "SLA Compliance:"
    ws['B5'] = f"{data['sla_compliance']}%"
    ws['A6'] = "First Call Resolution Rate:"
    ws['B6'] = f"{data['fcr_rate']}%"
    
    # Format numbers
    ws['B3'].number_format = '0'
    ws['B4'].number_format = '0.0'
    ws['B5'].number_format = '0.0%'
    ws['B6'].number_format = '0.0%'
    
    # Location breakdown
    ws['A8'] = "Location Breakdown"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Location"
    ws['B9'] = "Incidents"
    ws['C9'] = "MTTR (hours)"
    ws['D9'] = "SLA Compliance"
    ws['E9'] = "FCR Rate"
    
    style_header(ws, 9)
    
    row = 10
    for loc in data['locations']:
        ws[f'A{row}'] = loc['name']
        ws[f'B{row}'] = loc['incidents']
        ws[f'C{row}'] = loc['mttr']
        ws[f'D{row}'] = loc['sla_compliance'] / 100
        ws[f'E{row}'] = loc['fcr_rate'] / 100
        
        ws[f'B{row}'].number_format = '0'
        ws[f'C{row}'].number_format = '0.0'
        ws[f'D{row}'].number_format = '0.0%'
        ws[f'E{row}'].number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A9:E{row-1}')
    
    # Top technicians
    ws['A12'] = "Top Technicians"
    ws['A12'].font = Font(bold=True)
    
    ws['A13'] = "Technician"
    ws['B13'] = "Location"
    ws['C13'] = "Incidents"
    ws['D13'] = "% of Location Total"
    
    style_header(ws, 13)
    
    row = 14
    for tech in data['top_technicians']:
        ws[f'A{row}'] = tech['name']
        ws[f'B{row}'] = tech['location']
        ws[f'C{row}'] = tech['incidents']
        
        # Calculate percentage of location total
        location_data = next((loc for loc in data['locations'] if loc['name'] == tech['location']), None)
        if location_data and location_data['incidents'] > 0:
            percentage = (tech['incidents'] / location_data['incidents']) * 100
            ws[f'D{row}'] = percentage / 100
            ws[f'D{row}'].number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A13:D{row-1}')
    
    # Top KB articles
    ws['A17'] = "Top KB Articles"
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = "KB Article"
    ws['B18'] = "Description"
    ws['C18'] = "Usage Count"
    ws['D18'] = "% of Total KB Usage"
    
    style_header(ws, 18)
    
    row = 19
    total_kb_usage = sum(kb['count'] for kb in data['top_kb_articles'])
    
    for kb in data['top_kb_articles']:
        ws[f'A{row}'] = kb['kb_id']
        ws[f'B{row}'] = kb['description']
        ws[f'C{row}'] = kb['count']
        ws[f'D{row}'] = kb['count'] / total_kb_usage
        
        ws[f'D{row}'].number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A18:D{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    return ws

# Function to create a location-specific sheet
def create_location_sheet(wb, location_name, location_data):
    ws = wb.create_sheet(location_name)
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{location_name} - June 2025 Performance"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    ws['A3'] = "Total Incidents:"
    ws['B3'] = location_data['incidents']
    ws['A4'] = "Average MTTR (hours):"
    ws['B4'] = location_data['mttr']
    ws['A5'] = "SLA Compliance:"
    ws['B5'] = location_data['sla_compliance'] / 100
    ws['A6'] = "First Call Resolution Rate:"
    ws['B6'] = location_data['fcr_rate'] / 100
    
    # Format numbers
    ws['B3'].number_format = '0'
    ws['B4'].number_format = '0.0'
    ws['B5'].number_format = '0.0%'
    ws['B6'].number_format = '0.0%'
    
    # Technicians
    ws['A8'] = "Technicians"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Technician"
    ws['B9'] = "Incidents"
    ws['C9'] = "% of Location Total"
    ws['D9'] = "Avg MTTR (hours)"
    
    style_header(ws, 9)
    
    row = 10
    if 'technicians' in location_data and location_data['technicians']:
        for tech in location_data['technicians']:
            ws[f'A{row}'] = tech['name']
            ws[f'B{row}'] = tech['incidents']
            ws[f'C{row}'] = tech['percentage'] / 100
            
            # Add MTTR for each technician (using location average as placeholder)
            # In a real scenario, you would use actual technician MTTR data
            ws[f'D{row}'] = location_data['mttr']
            
            ws[f'B{row}'].number_format = '0'
            ws[f'C{row}'].number_format = '0.0%'
            ws[f'D{row}'].number_format = '0.0'
            
            row += 1
    else:
        ws[f'A{row}'] = "No technicians found for this location"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Add borders
    add_borders(ws, f'A9:D{row-1}')
    
    # KB Articles
    ws['A12'] = "KB Articles"
    ws['A12'].font = Font(bold=True)
    
    ws['A13'] = "KB Article"
    ws['B13'] = "Description"
    ws['C13'] = "Usage Count"
    ws['D13'] = "% of Location KB Usage"
    
    style_header(ws, 13)
    
    row = 14
    if 'kb_articles' in location_data and location_data['kb_articles']:
        total_location_kb = sum(kb['count'] for kb in location_data['kb_articles'])
        
        for kb in location_data['kb_articles']:
            ws[f'A{row}'] = kb['kb_id']
            ws[f'B{row}'] = kb['description']
            ws[f'C{row}'] = kb['count']
            ws[f'D{row}'] = kb['count'] / total_location_kb
            
            ws[f'D{row}'].number_format = '0.0%'
            
            row += 1
    else:
        ws[f'A{row}'] = "No KB articles found for this location"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Add borders
    add_borders(ws, f'A13:D{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    return ws

# Function to create the raw data sheet
def create_raw_data_sheet(wb, data):
    ws = wb.create_sheet("Raw Data")
    
    # Headers
    headers = ["Location", "Total Incidents", "MTTR (hours)", "SLA Compliance", "FCR Rate"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    style_header(ws)
    
    # Data
    row = 2
    for loc in data['locations']:
        ws.cell(row=row, column=1).value = loc['name']
        ws.cell(row=row, column=2).value = loc['incidents']
        ws.cell(row=row, column=3).value = loc['mttr']
        ws.cell(row=row, column=4).value = loc['sla_compliance'] / 100
        ws.cell(row=row, column=5).value = loc['fcr_rate'] / 100
        
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=5).number_format = '0.0%'
        
        row += 1
    
    # Add borders
    add_borders(ws, f'A1:E{row-1}')
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return ws

# Function to create the incident details sheet
def create_incident_details_sheet(wb, data):
    ws = wb.create_sheet("Incident Details")
    
    # Headers
    headers = ["Location", "Technician", "Incidents", "% of Location", "Avg MTTR (hours)", "KB Article", "Description", "Usage Count"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    style_header(ws)
    
    # Data
    row = 2
    for loc in data['locations']:
        if 'technicians' in loc and loc['technicians']:
            for tech in loc['technicians']:
                ws.cell(row=row, column=1).value = loc['name']
                ws.cell(row=row, column=2).value = tech['name']
                ws.cell(row=row, column=3).value = tech['incidents']
                ws.cell(row=row, column=4).value = tech['percentage'] / 100
                ws.cell(row=row, column=5).value = loc['mttr']  # Using location MTTR as placeholder
                
                ws.cell(row=row, column=4).number_format = '0.0%'
                ws.cell(row=row, column=5).number_format = '0.0'
                
                row += 1
        else:
            ws.cell(row=row, column=1).value = loc['name']
            ws.cell(row=row, column=2).value = "No technicians found"
            row += 1
    
    # Add borders
    add_borders(ws, f'A1:H{row-1}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 15
    
    return ws

# Main function to generate the report
def generate_central_report():
    # Central Region data
    central_data = {
        'total_incidents': 749,
        'avg_mttr': 1.3,
        'sla_compliance': 100.0,
        'fcr_rate': 99.6,
        'locations': [
            {
                'name': 'Homeoffice',
                'incidents': 221,
                'mttr': 1.2,
                'sla_compliance': 100.0,
                'fcr_rate': 100.0,
                'technicians': [
                    {'name': 'Jackie Phrakousonh', 'incidents': 98, 'percentage': 44.3},
                    {'name': 'Bryce Breedlove', 'incidents': 72, 'percentage': 32.6},
                    {'name': 'Stephanie Pham', 'incidents': 51, 'percentage': 23.1}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 32},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 28},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 25},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 21},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 18}
                ]
            },
            {
                'name': 'DGTC',
                'incidents': 231,
                'mttr': 1.5,
                'sla_compliance': 100.0,
                'fcr_rate': 99.1,
                'technicians': [
                    {'name': 'Agustin Rodriguez', 'incidents': 88, 'percentage': 38.1},
                    {'name': 'Mason Montgomery', 'incidents': 85, 'percentage': 36.8},
                    {'name': 'Elaine Nguyen', 'incidents': 58, 'percentage': 25.1}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 35},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 30},
                    {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 27},
                    {'kb_id': 'KB1148355 v3.0', 'description': 'Authentication & Access Issues', 'count': 22},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 19}
                ]
            },
            {
                'name': 'Jst',
                'incidents': 102,
                'mttr': 1.0,
                'sla_compliance': 100.0,
                'fcr_rate': 100.0,
                'technicians': [],  # Removed simulated names - needs real data
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 18},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 15},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 12},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 10},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 8}
                ]
            },
            {
                'name': "Sam's Club",
                'incidents': 98,
                'mttr': 1.4,
                'sla_compliance': 100.0,
                'fcr_rate': 99.0,
                'technicians': [],  # Removed simulated names - needs real data
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 22},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 17},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 15},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 9},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 7}
                ]
            },
            {
                'name': 'I Street',
                'incidents': 97,
                'mttr': 1.2,
                'sla_compliance': 100.0,
                'fcr_rate': 100.0,
                'technicians': [],  # Removed simulated names - needs real data
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 14},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 12},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 11},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 10},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 8}
                ]
            },
            {
                'name': 'Hula',
                'incidents': 162,
                'mttr': 1.8,
                'sla_compliance': 100.0,
                'fcr_rate': 99.4,
                'technicians': [
                    {'name': 'Plas Abraham', 'incidents': 61, 'percentage': 37.7},
                    {'name': 'Ben Masten', 'incidents': 59, 'percentage': 36.4},
                    {'name': 'Dillon Burch', 'incidents': 22, 'percentage': 13.6},
                    {'name': 'Harrison Couch', 'incidents': 13, 'percentage': 8.0},
                    {'name': 'Jon Lowe', 'incidents': 6, 'percentage': 3.7},
                    {'name': 'Hasheema Ali', 'incidents': 1, 'percentage': 0.6}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 25},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 22},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 20},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 18},
                    {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 15}
                ]
            },
            {
                'name': 'Purpose',
                'incidents': 108,
                'mttr': 0.9,
                'sla_compliance': 100.0,
                'fcr_rate': 100.0,
                'technicians': [
                    {'name': 'Tessa Black', 'incidents': 54, 'percentage': 50.0},
                    {'name': 'Kaleb Thompson', 'incidents': 53, 'percentage': 49.1},
                    {'name': 'Anthony Clark', 'incidents': 1, 'percentage': 0.9}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 19},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 17},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 14},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 12},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 9}
                ]
            },
            {
                'name': 'MLK',
                'incidents': 0,
                'mttr': 0.0,
                'sla_compliance': 0.0,
                'fcr_rate': 0.0,
                'technicians': [],
                'kb_articles': []
            },
            {
                'name': 'Aviation',
                'incidents': 0,
                'mttr': 0.0,
                'sla_compliance': 0.0,
                'fcr_rate': 0.0,
                'technicians': [],
                'kb_articles': []
            }
        ],
        'top_technicians': [
            {'name': 'Jackie Phrakousonh', 'location': 'Homeoffice', 'incidents': 98},
            {'name': 'Agustin Rodriguez', 'location': 'DGTC', 'incidents': 88},
            {'name': 'Mason Montgomery', 'location': 'DGTC', 'incidents': 85},
            {'name': 'Bryce Breedlove', 'location': 'Homeoffice', 'incidents': 72},
            {'name': 'Plas Abraham', 'location': 'Hula', 'incidents': 61},
            {'name': 'Ben Masten', 'location': 'Hula', 'incidents': 59},
            {'name': 'Elaine Nguyen', 'location': 'DGTC', 'incidents': 58},
            {'name': 'Tessa Black', 'location': 'Purpose', 'incidents': 54},
            {'name': 'Kaleb Thompson', 'location': 'Purpose', 'incidents': 53},
            {'name': 'Stephanie Pham', 'location': 'Homeoffice', 'incidents': 51}
        ],
        'top_kb_articles': [
            {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 103},
            {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 98},
            {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 77},
            {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 76},
            {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 52},
            {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 42},
            {'kb_id': 'KB1148355 v3.0', 'description': 'Authentication & Access Issues', 'count': 22}
        ]
    }
    
    # Create the workbook
    wb = create_excel_report("Updated_Central_Region_June_2025_Report.xlsx")
    
    # Create sheets
    create_executive_summary(wb, central_data)
    
    for location in central_data['locations']:
        create_location_sheet(wb, location['name'], location)
    
    create_raw_data_sheet(wb, central_data)
    create_incident_details_sheet(wb, central_data)
    
    # Save the workbook
    wb.save("Updated_Central_Region_June_2025_Report.xlsx")
    print("Updated Central Region report generated successfully!")

if __name__ == "__main__":
    generate_central_report() 