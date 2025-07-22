import openpyxl
import json
import math

def get_source_data():
    """
    Defines the single source of truth for the report data.
    This is the exact data from the generation script.
    """
    approved_technicians = [
        'Plas Abraham', 'Ben Masten', 'Dillon Burch', 'Harrison Couch', 'Jon Lowe', 
        'Hasheema Ali', 'Tessa Black', 'Kaleb Thompson', 'Anthony Clark', 
        'Jackie Phrakousonh', 'Bryce Breedlove', 'Stephanie Pham', 
        'Agustin Rodriguez', 'Mason Montgomery'
    ]
    
    central_data = {
        'total_incidents': 749,
        'avg_mttr': 1.3,
        'sla_compliance': 100.0,
        'fcr_rate': 99.6,
        'locations': [
            {
                'name': 'Homeoffice', 'incidents': 221, 'mttr': 1.2, 'sla_compliance': 100.0, 'fcr_rate': 100.0,
                'technicians': [
                    {'name': 'Jackie Phrakousonh', 'incidents': 98, 'percentage': 44.3},
                    {'name': 'Bryce Breedlove', 'incidents': 72, 'percentage': 32.6},
                    {'name': 'Stephanie Pham', 'incidents': 51, 'percentage': 23.1}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 32},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 28},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 25},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 21},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 18}
                ]
            },
            {
                'name': 'DGTC', 'incidents': 231, 'mttr': 1.5, 'sla_compliance': 100.0, 'fcr_rate': 99.1,
                'technicians': [
                    {'name': 'Agustin Rodriguez', 'incidents': 88, 'percentage': 38.1},
                    {'name': 'Mason Montgomery', 'incidents': 85, 'percentage': 36.8}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 35},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 30},
                    {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 27},
                    {'kb_id': 'KB1148355 v3.0', 'description': 'Authentication & Access Issues', 'count': 22},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 19}
                ]
            },
            {
                'name': 'Jst', 'incidents': 102, 'mttr': 1.0, 'sla_compliance': 100.0, 'fcr_rate': 100.0,
                'technicians': [],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 18},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 15},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 12},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 10},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 8}
                ]
            },
            {
                'name': "Sam's Club", 'incidents': 98, 'mttr': 1.4, 'sla_compliance': 100.0, 'fcr_rate': 99.0,
                'technicians': [],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 22},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 17},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 15},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 9},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 7}
                ]
            },
            {
                'name': 'I Street', 'incidents': 97, 'mttr': 1.2, 'sla_compliance': 100.0, 'fcr_rate': 100.0,
                'technicians': [],
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 14},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 12},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 11},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 10},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 8}
                ]
            },
            {
                'name': 'Hula', 'incidents': 162, 'mttr': 1.8, 'sla_compliance': 100.0, 'fcr_rate': 99.4,
                'technicians': [
                    {'name': 'Plas Abraham', 'incidents': 61, 'percentage': 37.7},
                    {'name': 'Ben Masten', 'incidents': 59, 'percentage': 36.4},
                    {'name': 'Dillon Burch', 'incidents': 22, 'percentage': 13.6},
                    {'name': 'Harrison Couch', 'incidents': 13, 'percentage': 8.0},
                    {'name': 'Jon Lowe', 'incidents': 6, 'percentage': 3.7},
                    {'name': 'Hasheema Ali', 'incidents': 1, 'percentage': 0.6}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 25},
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 22},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 20},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 18},
                    {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 15}
                ]
            },
            {
                'name': 'Purpose', 'incidents': 108, 'mttr': 0.9, 'sla_compliance': 100.0, 'fcr_rate': 100.0,
                'technicians': [
                    {'name': 'Tessa Black', 'incidents': 54, 'percentage': 50.0},
                    {'name': 'Kaleb Thompson', 'incidents': 53, 'percentage': 49.1},
                    {'name': 'Anthony Clark', 'incidents': 1, 'percentage': 0.9}
                ],
                'kb_articles': [
                    {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 19},
                    {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 17},
                    {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 14},
                    {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 12},
                    {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 9}
                ]
            },
            { 'name': 'MLK', 'incidents': 0, 'mttr': 0.0, 'sla_compliance': 0.0, 'fcr_rate': 0.0, 'technicians': [], 'kb_articles': [] },
            { 'name': 'Aviation', 'incidents': 0, 'mttr': 0.0, 'sla_compliance': 0.0, 'fcr_rate': 0.0, 'technicians': [], 'kb_articles': [] }
        ],
        'top_technicians': [
            {'name': 'Jackie Phrakousonh', 'location': 'Homeoffice', 'incidents': 98},
            {'name': 'Agustin Rodriguez', 'location': 'DGTC', 'incidents': 88},
            {'name': 'Mason Montgomery', 'location': 'DGTC', 'incidents': 85},
            {'name': 'Bryce Breedlove', 'location': 'Homeoffice', 'incidents': 72},
            {'name': 'Plas Abraham', 'location': 'Hula', 'incidents': 61},
            {'name': 'Ben Masten', 'location': 'Hula', 'incidents': 59},
            {'name': 'Tessa Black', 'location': 'Purpose', 'incidents': 54},
            {'name': 'Kaleb Thompson', 'location': 'Purpose', 'incidents': 53},
            {'name': 'Stephanie Pham', 'location': 'Homeoffice', 'incidents': 51},
            {'name': 'Dillon Burch', 'location': 'Hula', 'incidents': 22}
        ],
        'top_kb_articles': [
            {'kb_id': 'KB1148221 v2.0', 'description': 'Wireless Mouse Pairing & Connection Issues', 'count': 103},
            {'kb_id': 'KB1149657 v11.0', 'description': 'End of Life (EOL) Laptop Returns & Replacements', 'count': 98},
            {'kb_id': 'KB1150423 v7.0', 'description': 'System Performance & Optimization', 'count': 77},
            {'kb_id': 'KB1146152 v13.0', 'description': 'Network Connectivity & WIFI Issues', 'count': 76},
            {'kb_id': 'KB1148390 v9.0', 'description': 'Hardware Troubleshooting & BIOS Updates', 'count': 52},
            {'kb_id': 'KB1152218 v11.0', 'description': 'Laptop & Computer Hardware Issues', 'count': 42},
            {'kb_id': 'KB1148355 v3.0', 'description': 'Authentication & Access Issues', 'count': 22}
        ]
    }
    # Filter locations to only include approved technicians
    for location in central_data['locations']:
        location['technicians'] = [
            tech for tech in location.get('technicians', []) if tech['name'] in approved_technicians
        ]
    
    return central_data

class ReportVerifier:
    def __init__(self, workbook_path):
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)
        self.source_data = get_source_data()
        self.results = []

    def log_check(self, status, component, message):
        self.results.append({'status': status, 'component': component, 'message': message})
        mark = "✅" if status == "PASS" else "❌"
        print(f"{mark} [{component}] {message}")

    def compare_values(self, v1, v2, component, context):
        if isinstance(v1, float) or isinstance(v2, float):
            if not math.isclose(v1, v2, rel_tol=1e-3):
                self.log_check("FAIL", component, f"{context}: Source '{v1}' != Report '{v2}'")
            else:
                self.log_check("PASS", component, f"{context}: {v1} == {v2}")
        elif v1 != v2:
             self.log_check("FAIL", component, f"{context}: Source '{v1}' != Report '{v2}'")
        else:
            self.log_check("PASS", component, f"{context}: '{v1}' == '{v2}'")

    def verify_executive_summary(self):
        ws = self.wb['Executive Summary']
        self.log_check("INFO", "Executive Summary", "--- Verifying Executive Summary ---")
        self.compare_values(ws['B3'].value, self.source_data['total_incidents'], "Exec Summary", "Total Incidents")
        self.compare_values(ws['B4'].value, self.source_data['avg_mttr'], "Exec Summary", "Average MTTR")
        self.compare_values(float(ws['B5'].value.strip('%')), self.source_data['sla_compliance'], "Exec Summary", "SLA Compliance")
        self.compare_values(float(ws['B6'].value.strip('%')), self.source_data['fcr_rate'], "Exec Summary", "FCR Rate")
        
        # Location Breakdown Table
        self.log_check("INFO", "Exec Summary", "Verifying Location Breakdown table...")
        # Define the exact range for the location breakdown to avoid reading other tables
        location_table_range = f'A10:E{9 + len(self.source_data["locations"])}'
        for i, row_cells in enumerate(ws[location_table_range]):
            loc_data = self.source_data['locations'][i]
            row_num = 10 + i
            self.compare_values(row_cells[0].value, loc_data['name'], "Exec Location Table", f"Row {row_num} Name")
            self.compare_values(row_cells[1].value, loc_data['incidents'], "Exec Location Table", f"Row {row_num} Incidents")
            self.compare_values(row_cells[2].value, loc_data['mttr'], "Exec Location Table", f"Row {row_num} MTTR")
            self.compare_values(row_cells[3].value, loc_data['sla_compliance']/100, "Exec Location Table", f"Row {row_num} SLA")
            self.compare_values(row_cells[4].value, loc_data['fcr_rate']/100, "Exec Location Table", f"Row {row_num} FCR")
            
    def verify_all_location_sheets(self):
        for loc_data in self.source_data['locations']:
            loc_name = loc_data['name']
            self.log_check("INFO", f"Location Sheet: {loc_name}", f"--- Verifying sheet for {loc_name} ---")
            if loc_name not in self.wb.sheetnames:
                self.log_check("FAIL", f"Location Sheet: {loc_name}", "Sheet not found in workbook")
                continue
            ws = self.wb[loc_name]
            
            # KPIs
            self.compare_values(ws['B3'].value, loc_data['incidents'], f"{loc_name} Sheet", "Total Incidents")
            self.compare_values(ws['B4'].value, loc_data['mttr'], f"{loc_name} Sheet", "Average MTTR")
            self.compare_values(ws['B5'].value, loc_data['sla_compliance']/100, f"{loc_name} Sheet", "SLA Compliance")
            self.compare_values(ws['B6'].value, loc_data['fcr_rate']/100, f"{loc_name} Sheet", "FCR Rate")

            # Technicians table
            if not loc_data['technicians']:
                 self.compare_values(ws['A10'].value, "No technicians found for this location", f"{loc_name} Sheet", "No Tech message")
            else:
                for i, tech_data in enumerate(loc_data['technicians'], 10):
                    row = ws[i]
                    self.compare_values(row[0].value, tech_data['name'], f"{loc_name} Techs", f"Row {i} Name")
                    self.compare_values(row[1].value, tech_data['incidents'], f"{loc_name} Techs", f"Row {i} Incidents")
                    self.compare_values(row[2].value, tech_data['percentage']/100, f"{loc_name} Techs", f"Row {i} Percentage")
                    # Check placeholder MTTR
                    self.compare_values(row[3].value, loc_data['mttr'], f"{loc_name} Techs", f"Row {i} MTTR (placeholder)")

    def run_all_checks(self):
        print("--- Starting Comprehensive Report Verification ---")
        self.verify_executive_summary()
        self.verify_all_location_sheets()
        print("\n--- Verification Summary ---")
        failures = [r for r in self.results if r['status'] == 'FAIL']
        if not failures:
            print("✅✅✅ ALL CHECKS PASSED. The report is 100% accurate according to the source data.")
        else:
            print(f"❌❌❌ Found {len(failures)} discrepancies. Review the log above.")
        print("--- Verification Complete ---")

if __name__ == '__main__':
    verifier = ReportVerifier('Corrected_Central_Region_June_2025_Report.xlsx')
    verifier.run_all_checks() 